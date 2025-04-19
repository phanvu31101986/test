import json
from openpyxl import Workbook
from openpyxl.styles import Alignment

FILENAME = "students.json"

def xuat_excel(students):
    if not students:
        print("Danh sách trống. Không có gì để xuất.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "DanhSachSinhVien"
    ws.append(["STT", "Họ tên", "Tuổi", "Điểm"])

    for i, sv in enumerate(students, start=1):
        ws.append([i, sv["name"], sv["age"], sv["score"]])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    filename = "students.xlsx"
    wb.save(filename)
    print(f"Đã xuất danh sách ra file '{filename}'")


def load_data():
    try:
        with open(FILENAME, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return []

def save_data(students):
    with open(FILENAME, "w", encoding="utf-8") as f:
        json.dump(students, f, ensure_ascii=False, indent=4)

def them_sinh_vien(students):
    name = input("Nhập tên: ")
    age = int(input("Nhập tuổi: "))
    score = float(input("Nhập điểm: "))
    students.append({"name": name, "age": age, "score": score})
    print("Đã thêm sinh viên.")

def hien_thi_danh_sach(students):
    if not students:
        print("Danh sách trống.")
        return
    print("\n DANH SÁCH SINH VIÊN:")
    for i, sv in enumerate(students, start=1):
        print(f"{i}. {sv['name']} - Tuổi: {sv['age']} - Điểm: {sv['score']}")

def tim_kiem_theo_ten(students):
    keyword = input("Nhập tên cần tìm: ").lower()
    found = [sv for sv in students if keyword in sv['name'].lower()]
    if found:
        print("\n KẾT QUẢ TÌM KIẾM:")
        for sv in found:
            print(f"{sv['name']} - Tuổi: {sv['age']} - Điểm: {sv['score']}")
    else:
        print("Không tìm thấy sinh viên.")

def diem_cao_nhat(students):
    if not students:
        print("Danh sách trống.")
        return
    max_score = max(sv["score"] for sv in students)
    top_students = [sv for sv in students if sv["score"] == max_score]
    print("\n SINH VIÊN CÓ ĐIỂM CAO NHẤT:")
    for sv in top_students:
        print(f"{sv['name']} - Tuổi: {sv['age']} - Điểm: {sv['score']}")

def sua_sinh_vien(students):
    hien_thi_danh_sach(students)
    try:
        index = int(input("Nhập số thứ tự sinh viên cần sửa: ")) - 1
        if 0 <= index < len(students):
            name = input("Tên mới (để trống nếu không đổi): ")
            age = input("Tuổi mới (để trống nếu không đổi): ")
            score = input("Điểm mới (để trống nếu không đổi): ")
            if name:
                students[index]["name"] = name
            if age:
                students[index]["age"] = int(age)
            if score:
                students[index]["score"] = float(score)
            print("Đã cập nhật thông tin sinh viên.")
        else:
            print("Số thứ tự không hợp lệ.")
    except ValueError:
        print("Nhập sai định dạng.")

def xoa_sinh_vien(students):
    hien_thi_danh_sach(students)
    try:
        index = int(input("Nhập số thứ tự sinh viên cần xóa: ")) - 1
        if 0 <= index < len(students):
            removed = students.pop(index)
            print(f"Đã xóa sinh viên: {removed['name']}")
        else:
            print("Số thứ tự không hợp lệ.")
    except ValueError:
        print("Nhập sai định dạng.")

def loc_theo_diem(students):
    try:
        min_score = float(input("Nhập điểm tối thiểu: "))
        filtered = [sv for sv in students if sv["score"] >= min_score]
        if filtered:
            print(f"\n Danh sách sinh viên có điểm >= {min_score}:")
            for sv in filtered:
                print(f"{sv['name']} - Tuổi: {sv['age']} - Điểm: {sv['score']}")
        else:
            print("Không có sinh viên nào đạt yêu cầu.")
    except ValueError:
        print("Nhập sai định dạng điểm.")

def menu():
    print("\n========= MENU =========")
    print("1. Thêm sinh viên")
    print("2. Hiển thị danh sách")
    print("3. Tìm kiếm theo tên")
    print("4. Hiển thị sinh viên điểm cao nhất")
    print("5. Sửa sinh viên")
    print("6. Xóa sinh viên")
    print("7. Lọc sinh viên theo điểm")
    print("8. Lưu danh sách vào file (JSON)")
    print("9. Xuất danh sách ra Excel (.xlsx)")
    print("0. Thoát")
    print("========================")

def main():
    students = load_data()
    while True:
        menu()
        choice = input("Chọn chức năng (0-9): ")
        if choice == "1":
            them_sinh_vien(students)
        elif choice == "2":
            hien_thi_danh_sach(students)
        elif choice == "3":
            tim_kiem_theo_ten(students)
        elif choice == "4":
            diem_cao_nhat(students)
        elif choice == "5":
            sua_sinh_vien(students)
        elif choice == "6":
            xoa_sinh_vien(students)
        elif choice == "7":
            loc_theo_diem(students)
        elif choice == "8":
            save_data(students)
            print("Dữ liệu đã được lưu.")
        elif choice == "9":
            xuat_excel(students)
        elif choice == "0":
            save_data(students)
            print("Tạm biệt!")
            break
        else:
            print("Lựa chọn không hợp lệ. Vui lòng chọn lại.")

if __name__ == "__main__":
    main()
